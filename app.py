# app.py
import streamlit as st
import pandas as pd
import numpy as np
import requests
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote
import re
from collections import Counter, defaultdict
import json
from tqdm import tqdm  # Note: tqdm may not display in Streamlit; using st.progress instead
import sys
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from tenacity import retry, stop_after_attempt, wait_exponential
import asyncio
import nest_asyncio
nest_asyncio.apply()

# Additional imports for specific libs
from habanero import Crossref
import aiohttp

# --- Глобальные настройки ---
EMAIL = st.secrets.get("EMAIL", "your.email@example.com")  # Use Streamlit secrets for email
MAX_WORKERS = 5
RETRIES = 3
DELAYS = [0.2, 0.5, 0.7, 1.0, 1.3, 1.5, 2.0]

# Кэши для данных
crossref_cache = {}
openalex_cache = {}
unified_cache = {}
citing_cache = defaultdict(list)
institution_cache = {}
journal_cache = {}

# --- Rate Limiter ---
class RateLimiter:
    def __init__(self, calls_per_second=5):
        self.calls_per_second = calls_per_second
        self.timestamps = []
        self.lock = threading.Lock()
    
    def wait_if_needed(self):
        with self.lock:
            now = time.time()
            self.timestamps = [ts for ts in self.timestamps if now - ts < 1.0]
            
            if len(self.timestamps) >= self.calls_per_second:
                sleep_time = 1.0 - (now - self.timestamps[0])
                if sleep_time > 0:
                    time.sleep(sleep_time)
                self.timestamps = self.timestamps[1:]
            
            self.timestamps.append(now)

rate_limiter = RateLimiter(calls_per_second=8)

# --- Адаптивная задержка ---
class AdaptiveDelayer:
    def __init__(self):
        self.lock = threading.Lock()
        self.delay_index = 0

    def wait(self, success=True):
        with self.lock:
            if success:
                self.delay_index = 0
            else:
                self.delay_index = min(self.delay_index + 1, len(DELAYS) - 1)
            delay = DELAYS[self.delay_index]
            time.sleep(delay)
            return delay

delayer = AdaptiveDelayer()

# --- Конфигурация ---
class JournalAnalyzerConfig:
    """Централизованная конфигурация"""
    def __init__(self):
        self.email = EMAIL
        self.max_workers = MAX_WORKERS
        self.retries = RETRIES
        self.delays = DELAYS
        self.timeouts = {
            'crossref': 15,
            'openalex': 10,
            'batch': 30
        }
        self.batch_sizes = {
            'metadata': 10,
            'citations': 5
        }

config = JournalAnalyzerConfig()

# --- Валидация и парсинг периода ---
def parse_period(period_str):
    years = set()
    parts = [p.strip() for p in period_str.replace(' ', '').split(',') if p.strip()]
    for part in parts:
        if '-' in part:
            try:
                s, e = map(int, part.split('-'))
                if 1900 <= s <= 2100 and 1900 <= e <= 2100 and s <= e:
                    years.update(range(s, e + 1))
                else:
                    st.error(f"⚠️ Диапазон вне 1900–2100 или некорректный: {part}")
            except ValueError:
                st.error(f"⚠️ Ошибка парсинга диапазона: {part}")
        else:
            try:
                y = int(part)
                if 1900 <= y <= 2100:
                    years.add(y)
                else:
                    st.error(f"⚠️ Год вне 1900–2100: {y}")
            except ValueError:
                st.error(f"⚠️ Не год: {part}")
    if not years:
        st.error("❌ Нет корректных годов.")
        return []
    return sorted(years)

# --- Валидация данных ---
def validate_and_clean_data(items):
    """Валидация и очистка полученных данных"""
    validated = []
    skipped_count = 0
    
    for item in items:
        if not item.get('DOI'):
            skipped_count += 1
            continue
            
        doi = item['DOI'].lower().strip()
        if not doi.startswith('10.'):
            skipped_count += 1
            continue
            
        date_parts = item.get('created', {}).get('date-parts', [[]])[0]
        if not date_parts or date_parts[0] < 1900:
            skipped_count += 1
            continue
            
        item['DOI'] = doi
        validated.append(item)
    
    if skipped_count > 0:
        st.warning(f"⚠️ Пропущено {skipped_count} статей из-за проблем с данными")
    st.success(f"✅ Валидировано: {len(validated)}/{len(items)} статей")
    return validated

# === 1. Название журнала ===
def get_journal_name(issn):
    if issn in crossref_cache.get('journals', {}):
        return crossref_cache['journals'][issn]
    url = f"https://api.openalex.org/sources?filter=issn:{issn}"
    for _ in range(RETRIES):
        try:
            rate_limiter.wait_if_needed()
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data['meta']['count'] > 0:
                    name = data['results'][0]['display_name']
                    crossref_cache.setdefault('journals', {})[issn] = name
                    delayer.wait(success=True)
                    return name
        except:
            pass
        delayer.wait(success=False)
    return "Журнал не найден"

# === 2. Получение журнала по ISSN ===
def get_journal_by_issn(issn):
    """Получение информации о журнале по ISSN"""
    if issn in journal_cache:
        return journal_cache[issn]
    
    url = f"https://api.openalex.org/sources?filter=issn:{issn}"
    for _ in range(RETRIES):
        try:
            rate_limiter.wait_if_needed()
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data['meta']['count'] > 0:
                    journal_data = data['results'][0]
                    journal_cache[issn] = journal_data
                    delayer.wait(success=True)
                    return journal_data
        except:
            pass
        delayer.wait(success=False)
    return None

# === 3. Получение статей из Crossref ===
def fetch_articles_by_issn_period(issn, from_date, until_date):
    base_url = "https://api.crossref.org/works"
    items = []
    cursor = "*"
    params = {
        'filter': f'issn:{issn},from-pub-date:{from_date},until-pub-date:{until_date}',
        'rows': 1000,
        'cursor': cursor,
        'mailto': EMAIL
    }
    
    with st.progress(0) as progress_bar:
        current_progress = 0
        max_steps = 100  # Arbitrary for progress
        st.info("📥 Загрузка статей из Crossref...")
        
        while cursor:
            params['cursor'] = cursor
            success = False
            for _ in range(RETRIES):
                try:
                    rate_limiter.wait_if_needed()
                    resp = requests.get(base_url, params=params, timeout=15)
                    if resp.status_code == 200:
                        data = resp.json()
                        new_items = data['message']['items']
                        items.extend(new_items)
                        cursor = data['message'].get('next-cursor')
                        current_progress += len(new_items) / max_steps * 100  # Approximate
                        progress_bar.progress(min(current_progress, 100))
                        st.caption(f"Получено {len(items)} статей, курсор: {cursor[:10]}...")
                        delayer.wait(success=True)
                        success = True
                        break
                except Exception as e:
                    st.error(f"   Ошибка: {e}")
                delayer.wait(success=False)
            if not success:
                break
            if not new_items:
                break
    return items

# === 4. Получение Crossref metadata (кэшировано) ===
@retry(stop=stop_after_attempt(RETRIES), wait=wait_exponential(multiplier=1, min=4, max=10))
def get_crossref_metadata(doi):
    if doi in crossref_cache:
        return crossref_cache[doi]
    if not doi or doi == 'N/A':
        return None
    url = f"https://api.crossref.org/works/{quote(doi)}"
    headers = {'User-Agent': f"YourApp/1.0 (mailto:{EMAIL})"}
    rate_limiter.wait_if_needed()
    resp = requests.get(url, headers=headers, timeout=15)
    if resp.status_code == 200:
        data = resp.json()['message']
        crossref_cache[doi] = data
        delayer.wait(success=True)
        return data
    delayer.wait(success=False)
    raise Exception("Failed to fetch Crossref metadata")

# === 5. Получение OpenAlex metadata (кэшировано) ===
@retry(stop=stop_after_attempt(RETRIES), wait=wait_exponential(multiplier=1, min=4, max=10))
def get_openalex_metadata(doi):
    if doi in openalex_cache:
        return openalex_cache[doi]
    if not doi or doi == 'N/A':
        return None
    normalized = doi if doi.startswith('http') else f"https://doi.org/{doi}"
    url = f"https://api.openalex.org/works/{quote(normalized)}"
    rate_limiter.wait_if_needed()
    resp = requests.get(url, timeout=10)
    if resp.status_code == 200:
        data = resp.json()
        openalex_cache[doi] = data
        delayer.wait(success=True)
        return data
    delayer.wait(success=False)
    raise Exception("Failed to fetch OpenAlex metadata")

# === 6. Извлечение аффилиаций и стран из OpenAlex данных ===
def extract_affiliations_and_countries(openalex_data):
    """Извлечение аффилиаций и стран из данных OpenAlex"""
    affiliations = set()
    countries = set()
    authors_list = []
    
    if not openalex_data:
        return authors_list, list(affiliations), list(countries)
    
    if 'authorships' in openalex_data:
        for auth in openalex_data['authorships']:
            author_name = auth.get('author', {}).get('display_name', 'Unknown')
            authors_list.append(author_name)
            
            for inst in auth.get('institutions', []):
                inst_name = inst.get('display_name')
                country_code = inst.get('country_code')
                
                if inst_name:
                    affiliations.add(inst_name)
                if country_code:
                    countries.add(country_code.upper())
    
    return authors_list, list(affiliations), list(countries)

# === 7. Извлечение информации о журнале из метаданных ===
def extract_journal_info(metadata):
    """Извлечение информации о журнале из метаданных"""
    journal_info = {
        'issn': [],
        'journal_name': '',
        'publisher': ''
    }
    
    if not metadata:
        return journal_info
    
    cr = metadata.get('crossref')
    if cr:
        journal_info['issn'] = cr.get('ISSN', [])
        journal_info['journal_name'] = cr.get('container-title', [''])[0] if cr.get('container-title') else ''
        journal_info['publisher'] = cr.get('publisher', '')
    
    oa = metadata.get('openalex')
    if oa:
        host_venue = oa.get('host_venue', {})
        if host_venue:
            if not journal_info['journal_name']:
                journal_info['journal_name'] = host_venue.get('display_name', '')
            if not journal_info['publisher']:
                journal_info['publisher'] = host_venue.get('publisher', '')
            if not journal_info['issn']:
                journal_info['issn'] = host_venue.get('issn', [])
    
    return journal_info

# === 8. Унифицированные метаданные ===
def get_unified_metadata(doi):
    """Единый запрос для получения метаданных из обоих источников"""
    if doi in unified_cache:
        return unified_cache[doi]
    
    if not doi or doi == 'N/A':
        return {'crossref': None, 'openalex': None}
    
    with ThreadPoolExecutor(max_workers=2) as executor:
        future_cr = executor.submit(get_crossref_metadata, doi)
        future_oa = executor.submit(get_openalex_metadata, doi)
        cr_data = future_cr.result()
        oa_data = future_oa.result()
    
    result = {'crossref': cr_data, 'openalex': oa_data}
    unified_cache[doi] = result
    return result

# === 9. Получение цитирующих DOI и их metadata (параллельно, кэшировано) ===
def get_citing_dois_and_metadata(analyzed_doi):
    if analyzed_doi in citing_cache:
        return citing_cache[analyzed_doi]
    citing_list = []
    oa_data = get_openalex_metadata(analyzed_doi)
    if not oa_data or oa_data.get('cited_by_count', 0) == 0:
        citing_cache[analyzed_doi] = citing_list
        return citing_list
    work_id = oa_data['id'].split('/')[-1]
    url = f"https://api.openalex.org/works?filter=cites:{work_id}&per-page=100"
    cursor = "*"
    
    while cursor:
        success = False
        for _ in range(RETRIES):
            try:
                rate_limiter.wait_if_needed()
                resp = requests.get(f"{url}&cursor={cursor}", timeout=15)
                if resp.status_code == 200:
                    data = resp.json()
                    for w in data.get('results', []):
                        c_doi = w.get('doi')
                        if c_doi:
                            if c_doi not in crossref_cache:
                                get_crossref_metadata(c_doi)
                            if c_doi not in openalex_cache:
                                get_openalex_metadata(c_doi)
                            citing_list.append({
                                'doi': c_doi,
                                'pub_date': w.get('publication_date'),
                                'crossref': crossref_cache.get(c_doi),
                                'openalex': openalex_cache.get(c_doi)
                            })
                    cursor = data['meta'].get('next_cursor')
                    delayer.wait(success=True)
                    success = True
                    break
            except:
                pass
            delayer.wait(success=False)
        if not success:
            break
    citing_cache[analyzed_doi] = citing_list
    return citing_list

# === 10. Извлечение префикса DOI (для самоцитирования) ===
def get_doi_prefix(doi):
    if not doi or doi == 'N/A':
        return ''
    return doi.split('/')[0] if '/' in doi else doi[:7]

# === 11. Обработка с прогресс-баром (адаптировано для Streamlit) ===
def process_with_progress(items, func, desc="Обработка", unit="элементов"):
    """Обработка с прогресс-баром в Streamlit"""
    results = []
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(func, item): item for item in items}
        
        completed = 0
        for future in as_completed(futures):
            try:
                results.append(future.result())
            except Exception as e:
                st.error(f"Ошибка в {desc}: {e}")
                results.append(None)
            completed += 1
            progress_bar.progress(completed / len(items))
            status_text.text(f"{desc}: {completed}/{len(items)} {unit}")
    
    return results

# === 12. Анализ пересечений между анализируемыми и цитирующими работами (обновлено) ===
def analyze_overlaps(analyzed_metadata, citing_metadata):
    """Анализ пересечений с деталями DOI и overlaps"""
    
    # Общий анализ множеств
    analyzed_authors_set = set()
    analyzed_affiliations_set = set()
    analyzed_countries_set = set()
    
    for meta in analyzed_metadata:
        if meta and meta.get('openalex'):
            _, affs, countries = extract_affiliations_and_countries(meta['openalex'])
            authors = [a.get('family', '') for a in meta.get('crossref', {}).get('author', [])] if meta.get('crossref') else []
            analyzed_authors_set.update(authors)
            analyzed_affiliations_set.update(affs)
            analyzed_countries_set.update(countries)
    
    citing_authors_set = set()
    citing_affiliations_set = set()
    citing_countries_set = set()
    
    for meta in citing_metadata:
        if meta and meta.get('openalex'):
            _, affs, countries = extract_affiliations_and_countries(meta['openalex'])
            authors = [a.get('family', '') for a in meta.get('crossref', {}).get('author', [])] if meta.get('crossref') else []
            citing_authors_set.update(authors)
            citing_affiliations_set.update(affs)
            citing_countries_set.update(countries)
    
    author_overlap = analyzed_authors_set.intersection(citing_authors_set)
    affiliation_overlap = analyzed_affiliations_set.intersection(citing_affiliations_set)
    country_overlap = analyzed_countries_set.intersection(citing_countries_set)
    
    author_overlap_pct = (len(author_overlap) / len(analyzed_authors_set) * 100) if analyzed_authors_set else 0
    affiliation_overlap_pct = (len(affiliation_overlap) / len(analyzed_affiliations_set) * 100) if analyzed_affiliations_set else 0
    country_overlap_pct = (len(country_overlap) / len(analyzed_countries_set) * 100) if analyzed_countries_set else 0
    
    # Детализированные пересечения: pairs
    overlap_pairs = []
    for analyzed in analyzed_metadata:
        if not analyzed or not analyzed.get('crossref') or not analyzed.get('openalex'):
            continue
        analyzed_doi = analyzed['crossref'].get('DOI')
        analyzed_authors = set([a.get('family', '') for a in analyzed['crossref'].get('author', [])])
        analyzed_affs, _, _ = extract_affiliations_and_countries(analyzed['openalex'])
        analyzed_affs_set = set(analyzed_affs)
        
        for citing in citing_metadata:
            if not citing or not citing.get('crossref') or not citing.get('openalex'):
                continue
            citing_doi = citing['crossref'].get('DOI')
            citing_authors = set([a.get('family', '') for a in citing['crossref'].get('author', [])])
            citing_affs, _, _ = extract_affiliations_and_countries(citing['openalex'])
            citing_affs_set = set(citing_affs)
            
            common_authors = analyzed_authors.intersection(citing_authors)
            common_affs = analyzed_affs_set.intersection(citing_affs_set)
            
            if common_authors or common_affs:
                overlap_pairs.append({
                    'analyzed_doi': analyzed_doi,
                    'citing_doi': citing_doi,
                    'common_authors': list(common_authors),
                    'common_affiliations': list(common_affs)
                })
    
    return {
        'author_overlap': list(author_overlap),
        'author_overlap_count': len(author_overlap),
        'author_overlap_pct': author_overlap_pct,
        'affiliation_overlap': list(affiliation_overlap),
        'affiliation_overlap_count': len(affiliation_overlap),
        'affiliation_overlap_pct': affiliation_overlap_pct,
        'country_overlap': list(country_overlap),
        'country_overlap_count': len(country_overlap),
        'country_overlap_pct': country_overlap_pct,
        'total_analyzed_authors': len(analyzed_authors_set),
        'total_citing_authors': len(citing_authors_set),
        'total_analyzed_affiliations': len(analyzed_affiliations_set),
        'total_citing_affiliations': len(citing_affiliations_set),
        'total_analyzed_countries': len(analyzed_countries_set),
        'total_citing_countries': len(citing_countries_set),
        'overlap_pairs': overlap_pairs
    }

# === 13. Анализ скорости накопления цитирований ===
def analyze_citation_accumulation(analyzed_metadata):
    """Анализ скорости накопления цитирований по годам"""
    
    accumulation_data = defaultdict(lambda: defaultdict(int))
    yearly_citations = defaultdict(int)
    
    for analyzed in analyzed_metadata:
        if analyzed and analyzed.get('crossref'):
            analyzed_doi = analyzed['crossref'].get('DOI')
            if not analyzed_doi:
                continue
                
            pub_year = analyzed['crossref'].get('published', {}).get('date-parts', [[0]])[0][0]
            if not pub_year:
                continue
            
            citings = get_citing_dois_and_metadata(analyzed_doi)
            
            for citing in citings:
                if citing.get('openalex'):
                    cite_year = citing['openalex'].get('publication_year', 0)
                    if cite_year >= pub_year:
                        yearly_citations[cite_year] += 1
                        years_since_pub = cite_year - pub_year
                        if years_since_pub >= 0:
                            for year in range(years_since_pub + 1):
                                accumulation_data[pub_year][year] += 1
    
    accumulation_curves = {}
    for pub_year, yearly_counts in accumulation_data.items():
        sorted_years = sorted(yearly_counts.keys())
        cumulative_counts = []
        current_total = 0
        for year in sorted_years:
            current_total += yearly_counts[year]
            cumulative_counts.append({
                'years_since_publication': year,
                'cumulative_citations': current_total
            })
        accumulation_curves[pub_year] = cumulative_counts
    
    yearly_stats = []
    for year in sorted(yearly_citations.keys()):
        yearly_stats.append({
            'year': year,
            'citations_count': yearly_citations[year]
        })
    
    return {
        'accumulation_curves': dict(accumulation_curves),
        'yearly_citations': yearly_stats,
        'total_years_covered': len(yearly_citations)
    }

# === 14. Обработка метаданных для статистики (УНИВЕРСАЛЬНАЯ) ===
def extract_stats_from_metadata(metadata_list, is_analyzed=True, journal_prefix=''):
    """Универсальная функция для извлечения статистики из метаданных"""
    total_refs = 0
    refs_with_doi = 0
    refs_without_doi = 0
    self_cites = 0
    ref_counts = []
    author_counts = []
    single_authors = 0
    multi_authors_gt10 = 0
    author_freq = Counter()
    pub_dates = []
    
    articles_with_10_citations = 0
    articles_with_50_citations = 0
    articles_with_100_citations = 0
    articles_with_200_citations = 0

    affiliations_freq = Counter()
    countries_freq = Counter()
    single_country_articles = 0
    multi_country_articles = 0
    no_country_articles = 0
    all_authors = []
    all_affiliations = []
    all_countries = []
    
    journal_freq = Counter()
    publisher_freq = Counter()

    for meta in metadata_list:
        if not meta:
            continue

        cr = meta.get('crossref')
        if cr:
            refs = cr.get('reference', [])
            total_refs += len(refs)
            for ref in refs:
                ref_doi = ref.get('DOI', '')
                if ref_doi:
                    refs_with_doi += 1
                    if get_doi_prefix(ref_doi) == journal_prefix:
                        self_cites += 1
                else:
                    refs_without_doi += 1
            ref_counts.append(len(refs))

            authors = cr.get('author', [])
            num_auth = len(authors)
            author_counts.append(num_auth)
            if num_auth == 1:
                single_authors += 1
            if num_auth > 10:
                multi_authors_gt10 += 1

            for auth in authors:
                family = auth.get('family', '').strip().title()
                given = auth.get('given', '').strip()
                initials = '.'.join([c + '.' for c in given if c.isupper()]) if given else ''
                if initials:
                    name = f"{family} {initials}"
                else:
                    name = family or 'Unknown'
                author_freq[name] += 1

            date_parts = cr.get('published', {}).get('date-parts', [[datetime.now().year]])[0]
            pub_date = datetime(date_parts[0], date_parts[1] if len(date_parts)>1 else 1, date_parts[2] if len(date_parts)>2 else 1)
            pub_dates.append(pub_date)
            
            journal_name = cr.get('container-title', [''])[0] if cr.get('container-title') else ''
            publisher = cr.get('publisher', '')
            if journal_name:
                journal_freq[journal_name] += 1
            if publisher:
                publisher_freq[publisher] += 1

        oa = meta.get('openalex')
        if oa:
            authors_list, affiliations_list, countries_list = extract_affiliations_and_countries(oa)
            
            all_authors.extend(authors_list)
            all_affiliations.extend(affiliations_list)
            all_countries.extend(countries_list)
            
            for aff in affiliations_list:
                affiliations_freq[aff] += 1
            for country in countries_list:
                countries_freq[country] += 1
            
            unique_countries = set(countries_list)
            if len(unique_countries) == 0:
                no_country_articles += 1
            elif len(unique_countries) == 1:
                single_country_articles += 1
            elif len(unique_countries) > 1:
                multi_country_articles += 1
            
            host_venue = oa.get('host_venue', {})
            if host_venue:
                journal_name = host_venue.get('display_name', '')
                publisher = host_venue.get('publisher', '')
                if journal_name and journal_name not in journal_freq:
                    journal_freq[journal_name] += 1
                if publisher and publisher not in publisher_freq:
                    publisher_freq[publisher] += 1
            
            if is_analyzed:
                citation_count = oa.get('cited_by_count', 0)
                if citation_count >= 10:
                    articles_with_10_citations += 1
                if citation_count >= 50:
                    articles_with_50_citations += 1
                if citation_count >= 100:
                    articles_with_100_citations += 1
                if citation_count >= 200:
                    articles_with_200_citations += 1

    n_items = len(metadata_list)

    refs_with_doi_pct = (refs_with_doi / total_refs * 100) if total_refs > 0 else 0
    refs_without_doi_pct = (refs_without_doi / total_refs * 100) if total_refs > 0 else 0
    self_cites_pct = (self_cites / total_refs * 100) if total_refs > 0 else 0

    ref_min = min(ref_counts) if ref_counts else 0
    ref_max = max(ref_counts) if ref_counts else 0
    ref_mean = sum(ref_counts)/n_items if n_items > 0 else 0
    ref_median = sorted(ref_counts)[n_items//2] if n_items > 0 else 0

    auth_min = min(author_counts) if author_counts else 0
    auth_max = max(author_counts) if author_counts else 0
    auth_mean = sum(author_counts)/n_items if n_items > 0 else 0
    auth_median = sorted(author_counts)[n_items//2] if n_items > 0 else 0

    all_authors_sorted = author_freq.most_common()

    all_affiliations_sorted = affiliations_freq.most_common()
    all_countries_sorted = countries_freq.most_common()
    
    single_country_pct = (single_country_articles / n_items * 100) if n_items > 0 else 0
    multi_country_pct = (multi_country_articles / n_items * 100) if n_items > 0 else 0
    no_country_pct = (no_country_articles / n_items * 100) if n_items > 0 else 0

    all_journals_sorted = journal_freq.most_common()
    all_publishers_sorted = publisher_freq.most_common()

    return {
        'n_items': n_items,
        'total_refs': total_refs,
        'refs_with_doi': refs_with_doi, 'refs_with_doi_pct': refs_with_doi_pct,
        'refs_without_doi': refs_without_doi, 'refs_without_doi_pct': refs_without_doi_pct,
        'self_cites': self_cites, 'self_cites_pct': self_cites_pct,
        'ref_min': ref_min, 'ref_max': ref_max, 'ref_mean': ref_mean, 'ref_median': ref_median,
        'auth_min': auth_min, 'auth_max': auth_max, 'auth_mean': auth_mean, 'auth_median': auth_median,
        'single_authors': single_authors,
        'multi_authors_gt10': multi_authors_gt10,
        'all_authors': all_authors_sorted,
        'pub_dates': pub_dates,
        'articles_with_10_citations': articles_with_10_citations,
        'articles_with_50_citations': articles_with_50_citations,
        'articles_with_100_citations': articles_with_100_citations,
        'articles_with_200_citations': articles_with_200_citations,
        'all_affiliations': all_affiliations_sorted,
        'all_countries': all_countries_sorted,
        'all_authors_list': all_authors,
        'all_affiliations_list': all_affiliations,
        'all_countries_list': all_countries,
        'single_country_articles': single_country_articles, 
        'single_country_pct': single_country_pct,
        'multi_country_articles': multi_country_articles, 
        'multi_country_pct': multi_country_pct,
        'no_country_articles': no_country_articles,
        'no_country_pct': no_country_pct,
        'total_affiliations_count': len(all_affiliations),
        'unique_affiliations_count': len(set(all_affiliations)),
        'unique_countries_count': len(set(all_countries)),
        'all_journals': all_journals_sorted,
        'all_publishers': all_publishers_sorted,
        'unique_journals_count': len(journal_freq),
        'unique_publishers_count': len(publisher_freq)
    }

# === 15. Расчет расширенной статистики ===
def enhanced_stats_calculation(analyzed_metadata, citing_metadata):
    """Расширенная статистика с H-index и сетями цитирования"""
    
    citation_network = defaultdict(list)
    citation_counts = []
    
    for analyzed in analyzed_metadata:
        if analyzed and analyzed.get('crossref'):
            analyzed_doi = analyzed['crossref'].get('DOI')
            if analyzed_doi:
                analyzed_year = analyzed['crossref'].get('published', {}).get('date-parts', [[0]])[0][0]
                citings = get_citing_dois_and_metadata(analyzed_doi)
                citation_counts.append(len(citings))
                
                for citing in citings:
                    if citing.get('openalex'):
                        citing_year = citing['openalex'].get('publication_year', 0)
                        citation_network[analyzed_year].append(citing_year)
    
    citation_counts.sort(reverse=True)
    h_index = 0
    for i, count in enumerate(citation_counts):
        if count >= i + 1:
            h_index = i + 1
        else:
            break
    
    return {
        'h_index': h_index,
        'citation_network': dict(citation_network),
        'avg_citations_per_article': sum(citation_counts) / len(citation_counts) if citation_counts else 0,
        'max_citations': max(citation_counts) if citation_counts else 0,
        'min_citations': min(citation_counts) if citation_counts else 0,
        'total_citations': sum(citation_counts),
        'articles_with_citations': len([c for c in citation_counts if c > 0]),
        'articles_without_citations': len([c for c in citation_counts if c == 0])
    }

# === 16. Расчет времени до первого цитирования (расширено на все лаги) ===
def calculate_citation_timing_stats(analyzed_metadata):
    """Расчет времени между публикацией и цитированиями (первые и все лаги)"""
    all_days_to_first_citation = []
    all_lags_days = []  # Все лаги
    
    for analyzed in analyzed_metadata:
        if analyzed and analyzed.get('crossref'):
            analyzed_doi = analyzed['crossref'].get('DOI')
            if not analyzed_doi:
                continue
                
            analyzed_date_parts = analyzed['crossref'].get('published', {}).get('date-parts', [[]])[0]
            if not analyzed_date_parts or len(analyzed_date_parts) < 1:
                continue
                
            analyzed_year = analyzed_date_parts[0]
            analyzed_month = analyzed_date_parts[1] if len(analyzed_date_parts) > 1 else 1
            analyzed_day = analyzed_date_parts[2] if len(analyzed_date_parts) > 2 else 1
            
            try:
                analyzed_date = datetime(analyzed_year, analyzed_month, analyzed_day)
            except:
                continue
            
            citings = get_citing_dois_and_metadata(analyzed_doi)
            citation_dates = []
            
            for citing in citings:
                if citing.get('pub_date'):
                    try:
                        cite_date = datetime.fromisoformat(citing['pub_date'].replace('Z', '+00:00'))
                        if (cite_date - analyzed_date).days >= 0:
                            citation_dates.append(cite_date)
                            all_lags_days.append((cite_date - analyzed_date).days)
                    except:
                        continue
            
            if citation_dates:
                first_citation_date = min(citation_dates)
                days_to_first_citation = (first_citation_date - analyzed_date).days
                if days_to_first_citation >= 0:
                    all_days_to_first_citation.append(days_to_first_citation)
    
    first_timing_stats = {}
    if all_days_to_first_citation:
        first_timing_stats = {
            'min_days_to_first_citation': min(all_days_to_first_citation),
            'max_days_to_first_citation': max(all_days_to_first_citation),
            'mean_days_to_first_citation': np.mean(all_days_to_first_citation),
            'median_days_to_first_citation': np.median(all_days_to_first_citation),
            'articles_with_citation_timing_data': len(all_days_to_first_citation)
        }
    else:
        first_timing_stats = {
            'min_days_to_first_citation': 0,
            'max_days_to_first_citation': 0,
            'mean_days_to_first_citation': 0,
            'median_days_to_first_citation': 0,
            'articles_with_citation_timing_data': 0
        }
    
    all_lags_stats = {}
    if all_lags_days:
        all_lags_stats = {
            'min_lag_days': min(all_lags_days),
            'max_lag_days': max(all_lags_days),
            'mean_lag_days': np.mean(all_lags_days),
            'median_lag_days': np.median(all_lags_days),
            'total_citation_lags': len(all_lags_days)
        }
    else:
        all_lags_stats = {
            'min_lag_days': 0,
            'max_lag_days': 0,
            'mean_lag_days': 0,
            'median_lag_days': 0,
            'total_citation_lags': 0
        }
    
    return {
        'first_citation': first_timing_stats,
        'all_lags': all_lags_stats
    }

# === 17. Расчет Impact Factor (ИСПРАВЛЕННЫЙ) ===
def calculate_impact_factor(analyzed_metadata, current_date):
    """Расчет Impact Factor по правилам Web of Science"""
    
    current_year = current_date.year
    citation_years = [current_year - 1, current_year]
    publication_years = [current_year - 3, current_year - 2]
    
    st.info(f"🔍 Расчет IF {current_year}:")
    st.info(f"   Публикации: {publication_years[0]}-{publication_years[1]}")
    st.info(f"   Цитирования: {citation_years[0]}-{citation_years[1]}")
    
    publications_count = 0
    publications_list = []
    
    for meta in analyzed_metadata:
        if meta and meta.get('crossref'):
            pub_year = meta['crossref'].get('published', {}).get('date-parts', [[0]])[0][0]
            if pub_year in publication_years:
                publications_count += 1
                publications_list.append({
                    'doi': meta['crossref'].get('DOI'),
                    'year': pub_year,
                    'title': meta['crossref'].get('title', [''])[0]
                })
    
    citations_count = 0
    citations_details = []
    
    for pub in publications_list:
        if pub['doi']:
            citings = get_citing_dois_and_metadata(pub['doi'])
            for citing in citings:
                if citing.get('openalex'):
                    cite_year = citing['openalex'].get('publication_year', 0)
                    if cite_year in citation_years:
                        citations_count += 1
                        citations_details.append({
                            'cited_doi': pub['doi'],
                            'cited_year': pub['year'],
                            'citing_doi': citing['doi'],
                            'citing_year': cite_year
                        })
    
    impact_factor = citations_count / publications_count if publications_count > 0 else 0.0
    
    return {
        'impact_factor': impact_factor,
        'citations_count': citations_count,
        'publications_count': publications_count,
        'citation_years': citation_years,
        'publication_years': publication_years,
        'publications_list': publications_list,
        'citations_details': citations_details
    }

# === 18. Расчет IF и дней (ОБНОВЛЕННЫЙ) ===
def calculate_if_and_days(analyzed_metadata, all_citing_metadata, current_date):
    """Объединенный расчет Impact Factor и времени цитирования"""
    
    if_data = calculate_impact_factor(analyzed_metadata, current_date)
    
    timing_stats = calculate_citation_timing_stats(analyzed_metadata)
    
    accumulation_stats = analyze_citation_accumulation(analyzed_metadata)
    
    return {
        'if_value': if_data['impact_factor'],
        'c_num': if_data['citations_count'],
        'p_den': if_data['publications_count'],
        'citation_years': if_data['citation_years'],
        'publication_years': if_data['publication_years'],
        'first_citation': timing_stats['first_citation'],
        'all_lags': timing_stats['all_lags'],
        'accumulation_curves': accumulation_stats['accumulation_curves'],
        'yearly_citations': accumulation_stats['yearly_citations'],
        'total_years_covered': accumulation_stats['total_years_covered']
    }

# === 19. Создание расширенного Excel отчета (ПОЛНОСТЬЮ ПЕРЕРАБОТАННЫЙ) ===
def create_enhanced_excel_report(analyzed_data, citing_data, analyzed_stats, citing_stats, enhanced_stats, if_days, overlap_stats, filename):
    """Создание расширенного Excel отчета с полной статистикой"""
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # === ЛИСТ 1: Анализируемые статьи ===
    analyzed_list = []
    for item in analyzed_data:
        if item and item.get('crossref'):
            cr = item['crossref']
            oa = item.get('openalex', {})
            authors_list, affiliations_list, countries_list = extract_affiliations_and_countries(oa)
            journal_info = extract_journal_info(item)
            
            analyzed_list.append({
                'DOI': cr.get('DOI', ''),
                'Название': cr.get('title', [''])[0] if cr.get('title') else 'Без названия',
                'Авторы_Crossref': '; '.join([f"{a.get('given', '')} {a.get('family', '')}" for a in cr.get('author', [])]),
                'Авторы_OpenAlex': '; '.join(authors_list),
                'Аффилиации': '; '.join(affiliations_list),
                'Страны': '; '.join(countries_list),
                'Год публикации': cr.get('published', {}).get('date-parts', [[0]])[0][0],
                'Журнал': journal_info['journal_name'],
                'Издатель': journal_info['publisher'],
                'ISSN': '; '.join(journal_info['issn']),
                'Количество ссылок': cr.get('reference-count', 0),
                'Цитирования Crossref': cr.get('is-referenced-by-count', 0),
                'Цитирования OpenAlex': oa.get('cited_by_count', 0),
                'Количество авторов': len(cr.get('author', [])),
                'Тип работы': cr.get('type', '')
            })
    
    if analyzed_list:
        ws = wb.create_sheet('Анализируемые_статьи')
        df = pd.DataFrame(analyzed_list)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # === ЛИСТ 2: Цитирующие работы ===
    citing_list = []
    for item in citing_data:
        if item and item.get('crossref'):
            cr = item['crossref']
            oa = item.get('openalex', {})
            authors_list, affiliations_list, countries_list = extract_affiliations_and_countries(oa)
            journal_info = extract_journal_info(item)
            
            citing_list.append({
                'DOI': cr.get('DOI', ''),
                'Название': cr.get('title', [''])[0] if cr.get('title') else 'Без названия',
                'Авторы_Crossref': '; '.join([f"{a.get('given', '')} {a.get('family', '')}" for a in cr.get('author', [])]),
                'Авторы_OpenAlex': '; '.join(authors_list),
                'Аффилиации': '; '.join(affiliations_list),
                'Страны': '; '.join(countries_list),
                'Год публикации': cr.get('published', {}).get('date-parts', [[0]])[0][0],
                'Журнал': journal_info['journal_name'],
                'Издатель': journal_info['publisher'],
                'ISSN': '; '.join(journal_info['issn']),
                'Количество ссылок': cr.get('reference-count', 0),
                'Цитирования Crossref': cr.get('is-referenced-by-count', 0),
                'Цитирования OpenAlex': oa.get('cited_by_count', 0),
                'Количество авторов': len(cr.get('author', [])),
                'Тип работы': cr.get('type', '')
            })
    
    if citing_list:
        ws = wb.create_sheet('Цитирующие_работы')
        df = pd.DataFrame(citing_list)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # === ЛИСТ 3: Статистика анализируемых статей ===
    analyzed_stats_data = {
        'Метрика': [
            'Всего статей', 
            'Общее количество ссылок', 
            'Ссылки с DOI', 'Количество ссылок с DOI', 'Процент ссылок с DOI',
            'Ссылки без DOI', 'Количество ссылок без DOI', 'Процент ссылок без DOI',
            'Самоцитирования', 'Количество самоцитирований', 'Процент самоцитирований',
            'Статьи с одним автором',
            'Статьи с >10 авторами', 
            'Минимальное число ссылок', 
            'Максимальное число ссылок', 
            'Среднее число ссылок',
            'Медиана ссылок', 
            'Минимальное число авторов',
            'Максимальное число авторов', 
            'Среднее число авторов',
            'Медиана авторов', 
            'Статьи из одной страны', 'Процент статей из одной страны',
            'Статьи из нескольких стран', 'Процент статей из нескольких стран',
            'Статьи без данных о странах', 'Процент статей без данных о странах',
            'Всего аффилиаций',
            'Уникальных аффилиаций', 
            'Уникальных стран',
            'Уникальных журналов',
            'Уникальных издателей',
            'Статьи с ≥10 цитированиями',
            'Статьи с ≥50 цитированиями',
            'Статьи с ≥100 цитированиями',
            'Статьи с ≥200 цитированиями'
        ],
        'Значение': [
            analyzed_stats['n_items'],
            analyzed_stats['total_refs'],
            analyzed_stats['refs_with_doi'], f"{analyzed_stats['refs_with_doi_pct']:.1f}%",
            analyzed_stats['refs_without_doi'], f"{analyzed_stats['refs_without_doi_pct']:.1f}%",
            analyzed_stats['self_cites'], f"{analyzed_stats['self_cites_pct']:.1f}%",
            analyzed_stats['single_authors'],
            analyzed_stats['multi_authors_gt10'],
            analyzed_stats['ref_min'],
            analyzed_stats['ref_max'],
            f"{analyzed_stats['ref_mean']:.1f}",
            analyzed_stats['ref_median'],
            analyzed_stats['auth_min'],
            analyzed_stats['auth_max'],
            f"{analyzed_stats['auth_mean']:.1f}",
            analyzed_stats['auth_median'],
            analyzed_stats['single_country_articles'], f"{analyzed_stats['single_country_pct']:.1f}%",
            analyzed_stats['multi_country_articles'], f"{analyzed_stats['multi_country_pct']:.1f}%",
            analyzed_stats['no_country_articles'], f"{analyzed_stats['no_country_pct']:.1f}%",
            analyzed_stats['total_affiliations_count'],
            analyzed_stats['unique_affiliations_count'],
            analyzed_stats['unique_countries_count'],
            analyzed_stats['unique_journals_count'],
            analyzed_stats['unique_publishers_count'],
            analyzed_stats['articles_with_10_citations'],
            analyzed_stats['articles_with_50_citations'],
            analyzed_stats['articles_with_100_citations'],
            analyzed_stats['articles_with_200_citations']
        ]
    }
    ws = wb.create_sheet('Статистика_анализируемых')
    df = pd.DataFrame(analyzed_stats_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 4: Статистика цитирующих статей ===
    citing_stats_data = {
        'Метрика': [
            'Всего цитирующих статей', 
            'Общее количество ссылок', 
            'Ссылки с DOI', 'Количество ссылок с DOI', 'Процент ссылок с DOI',
            'Ссылки без DOI', 'Количество ссылок без DOI', 'Процент ссылок без DOI',
            'Самоцитирования', 'Количество самоцитирований', 'Процент самоцитирований',
            'Статьи с одним автором',
            'Статьи с >10 авторами', 
            'Минимальное число ссылок', 
            'Максимальное число ссылок', 
            'Среднее число ссылок',
            'Медиана ссылок', 
            'Минимальное число авторов',
            'Максимальное число авторов', 
            'Среднее число авторов',
            'Медиана авторов', 
            'Статьи из одной страны', 'Процент статей из одной страны',
            'Статьи из нескольких стран', 'Процент статей из нескольких стран',
            'Статьи без данных о странах', 'Процент статей без данных о странах',
            'Всего аффилиаций',
            'Уникальных аффилиаций', 
            'Уникальных стран',
            'Уникальных журналов',
            'Уникальных издателей'
        ],
        'Значение': [
            citing_stats['n_items'],
            citing_stats['total_refs'],
            citing_stats['refs_with_doi'], f"{citing_stats['refs_with_doi_pct']:.1f}%",
            citing_stats['refs_without_doi'], f"{citing_stats['refs_without_doi_pct']:.1f}%",
            citing_stats['self_cites'], f"{citing_stats['self_cites_pct']:.1f}%",
            citing_stats['single_authors'],
            citing_stats['multi_authors_gt10'],
            citing_stats['ref_min'],
            citing_stats['ref_max'],
            f"{citing_stats['ref_mean']:.1f}",
            citing_stats['ref_median'],
            citing_stats['auth_min'],
            citing_stats['auth_max'],
            f"{citing_stats['auth_mean']:.1f}",
            citing_stats['auth_median'],
            citing_stats['single_country_articles'], f"{citing_stats['single_country_pct']:.1f}%",
            citing_stats['multi_country_articles'], f"{citing_stats['multi_country_pct']:.1f}%",
            citing_stats['no_country_articles'], f"{citing_stats['no_country_pct']:.1f}%",
            citing_stats['total_affiliations_count'],
            citing_stats['unique_affiliations_count'],
            citing_stats['unique_countries_count'],
            citing_stats['unique_journals_count'],
            citing_stats['unique_publishers_count']
        ]
    }
    ws = wb.create_sheet('Статистика_цитирующих')
    df = pd.DataFrame(citing_stats_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 5: Все авторы анализируемых ===
    all_authors_data = {
        'Автор': [author[0] for author in analyzed_stats['all_authors']],
        'Количество статей': [author[1] for author in analyzed_stats['all_authors']]
    }
    ws = wb.create_sheet('Все_авторы_анализируемые')
    df = pd.DataFrame(all_authors_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 6: Все авторы цитирующих ===
    all_citing_authors_data = {
        'Автор': [author[0] for author in citing_stats['all_authors']],
        'Количество статей': [author[1] for author in citing_stats['all_authors']]
    }
    ws = wb.create_sheet('Все_авторы_цитирующие')
    df = pd.DataFrame(all_citing_authors_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 7: Все аффилиации анализируемых ===
    all_affiliations_data = {
        'Аффилиация': [aff[0] for aff in analyzed_stats['all_affiliations']],
        'Количество упоминаний': [aff[1] for aff in analyzed_stats['all_affiliations']]
    }
    ws = wb.create_sheet('Все_аффилиации_анализируемые')
    df = pd.DataFrame(all_affiliations_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 8: Все аффилиации цитирующих ===
    all_citing_affiliations_data = {
        'Аффилиация': [aff[0] for aff in citing_stats['all_affiliations']],
        'Количество упоминаний': [aff[1] for aff in citing_stats['all_affiliations']]
    }
    ws = wb.create_sheet('Все_аффилиации_цитирующие')
    df = pd.DataFrame(all_citing_affiliations_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 9: Все страны анализируемых ===
    all_countries_data = {
        'Страна': [country[0] for country in analyzed_stats['all_countries']],
        'Количество упоминаний': [country[1] for country in analyzed_stats['all_countries']]
    }
    ws = wb.create_sheet('Все_страны_анализируемые')
    df = pd.DataFrame(all_countries_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 10: Все страны цитирующих ===
    all_citing_countries_data = {
        'Страна': [country[0] for country in citing_stats['all_countries']],
        'Количество упоминаний': [country[1] for country in citing_stats['all_countries']]
    }
    ws = wb.create_sheet('Все_страны_цитирующие')
    df = pd.DataFrame(all_citing_countries_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 11: Все журналы цитирующих ===
    all_citing_journals_data = {
        'Журнал': [journal[0] for journal in citing_stats['all_journals']],
        'Количество статей': [journal[1] for journal in citing_stats['all_journals']]
    }
    ws = wb.create_sheet('Все_журналы_цитирующие')
    df = pd.DataFrame(all_citing_journals_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 12: Все издатели цитирующих ===
    all_citing_publishers_data = {
        'Издатель': [publisher[0] for publisher in citing_stats['all_publishers']],
        'Количество статей': [publisher[1] for publisher in citing_stats['all_publishers']]
    }
    ws = wb.create_sheet('Все_издатели_цитирующие')
    df = pd.DataFrame(all_citing_publishers_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 13: Пересечения анализируемых и цитирующих ===
    overlap_data = {
        'Категория': [
            'Авторы', 'Аффилиации', 'Страны'
        ],
        'Пересечение (количество)': [
            overlap_stats['author_overlap_count'],
            overlap_stats['affiliation_overlap_count'],
            overlap_stats['country_overlap_count']
        ],
        'Пересечение (%)': [
            f"{overlap_stats['author_overlap_pct']:.1f}%",
            f"{overlap_stats['affiliation_overlap_pct']:.1f}%",
            f"{overlap_stats['country_overlap_pct']:.1f}%"
        ],
        'Всего анализируемых': [
            overlap_stats['total_analyzed_authors'],
            overlap_stats['total_analyzed_affiliations'],
            overlap_stats['total_analyzed_countries']
        ],
        'Всего цитирующих': [
            overlap_stats['total_citing_authors'],
            overlap_stats['total_citing_affiliations'],
            overlap_stats['total_citing_countries']
        ]
    }
    ws = wb.create_sheet('Пересечения_анализируемые_цитирующие')
    df = pd.DataFrame(overlap_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 14: Детали пересечений пар ===
    overlap_pairs_data = []
    for pair in overlap_stats['overlap_pairs']:
        overlap_pairs_data.append({
            'Analyzed_DOI': pair['analyzed_doi'],
            'Citing_DOI': pair['citing_doi'],
            'Common_Authors': '; '.join(pair['common_authors']),
            'Common_Affiliations': '; '.join(pair['common_affiliations'])
        })
    ws = wb.create_sheet('Пересечения_пары')
    df = pd.DataFrame(overlap_pairs_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 15: Расширенная статистика ===
    enhanced_stats_data = {
        'Метрика': [
            'H-index', 'Общее количество цитирований',
            'Среднее цитирований на статью', 'Максимальное цитирований',
            'Минимальное цитирований', 'Статьи с цитированиями',
            'Статьи без цитирований'
        ],
        'Значение': [
            enhanced_stats['h_index'],
            enhanced_stats['total_citations'],
            f"{enhanced_stats['avg_citations_per_article']:.1f}",
            enhanced_stats['max_citations'],
            enhanced_stats['min_citations'],
            enhanced_stats['articles_with_citations'],
            enhanced_stats['articles_without_citations']
        ]
    }
    ws = wb.create_sheet('Расширенная_статистика')
    df = pd.DataFrame(enhanced_stats_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 16: Impact Factor, время цитирования и лаги ===
    if_days_data = {
        'Метрика': [
            'Impact Factor', 'Числитель (цитирования)', 
            'Знаменатель (публикации)', 'Годы цитирований',
            'Годы публикаций', 
            'Min дни до первого цитирования',
            'Max дни до первого цитирования', 'Mean дни до первого цитирования',
            'Median дни до первого цитирования', 'Статьи с данными о первом цитировании',
            'Min лаг всех цитирований',
            'Max лаг всех цитирований', 'Mean лаг всех цитирований',
            'Median лаг всех цитирований', 'Всего лагов цитирований',
            'Всего лет покрыто данными о цитированиях'
        ],
        'Значение': [
            f"{if_days['if_value']:.4f}",
            if_days['c_num'],
            if_days['p_den'],
            f"{if_days['citation_years'][0]}-{if_days['citation_years'][1]}",
            f"{if_days['publication_years'][0]}-{if_days['publication_years'][1]}",
            if_days['first_citation']['min_days_to_first_citation'],
            if_days['first_citation']['max_days_to_first_citation'],
            f"{if_days['first_citation']['mean_days_to_first_citation']:.1f}",
            if_days['first_citation']['median_days_to_first_citation'],
            if_days['first_citation']['articles_with_citation_timing_data'],
            if_days['all_lags']['min_lag_days'],
            if_days['all_lags']['max_lag_days'],
            f"{if_days['all_lags']['mean_lag_days']:.1f}",
            if_days['all_lags']['median_lag_days'],
            if_days['all_lags']['total_citation_lags'],
            if_days['total_years_covered']
        ]
    }
    ws = wb.create_sheet('Impact_Factor_Лаги_цитирования')
    df = pd.DataFrame(if_days_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # === ЛИСТ 17: Цитирования по годам ===
    yearly_citations_data = []
    for yearly_stat in if_days['yearly_citations']:
        yearly_citations_data.append({
            'Год': yearly_stat['year'],
            'Количество цитирований': yearly_stat['citations_count']
        })
    
    if yearly_citations_data:
        ws = wb.create_sheet('Цитирования_по_годам')
        df = pd.DataFrame(yearly_citations_data)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # === ЛИСТ 18: Кривые накопления цитирований ===
    accumulation_data = []
    for pub_year, curve_data in if_days['accumulation_curves'].items():
        for data_point in curve_data:
            accumulation_data.append({
                'Год публикации': pub_year,
                'Лет после публикации': data_point['years_since_publication'],
                'Накопительные цитирования': data_point['cumulative_citations']
            })
    
    if accumulation_data:
        ws = wb.create_sheet('Кривые_накопления_цитирований')
        df = pd.DataFrame(accumulation_data)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # === ЛИСТ 19: Сеть цитирований ===
    citation_network_data = []
    for year, citing_years in enhanced_stats.get('citation_network', {}).items():
        year_counts = Counter(citing_years)
        for citing_year, count in year_counts.items():
            citation_network_data.append({
                'Год публикации': year,
                'Год цитирования': citing_year,
                'Количество цитирований': count
            })
    
    if citation_network_data:
        ws = wb.create_sheet('Сеть_цитирований')
        df = pd.DataFrame(citation_network_data)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# === 20. Основная функция анализа (адаптирована для Streamlit) ===
def analyze_journal(issn, period_str):
    global delayer
    delayer = AdaptiveDelayer()
    
    st.title("🚀 РЕЗУЛЬТАТЫ АНАЛИЗА ЖУРНАЛА")
    st.markdown("---")
    
    overall_progress = st.progress(0)
    overall_status = st.empty()
    
    steps = 8  # Updated for new steps
    current_step = 0
    
    # Парсинг периода
    overall_status.text("📅 Парсинг периода...")
    years = parse_period(period_str)
    if not years:
        return
    from_date = f"{min(years)}-01-01"
    until_date = f"{max(years)}-12-31"
    st.info(f"📅 Период: {from_date} → {until_date}")
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Название журнала
    overall_status.text("📖 Получение названия журнала...")
    journal_name = get_journal_name(issn)
    st.success(f"📖 Журнал: {journal_name} (ISSN: {issn})")
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Получение статей
    overall_status.text("📥 Загрузка статей из Crossref...")
    items = fetch_articles_by_issn_period(issn, from_date, until_date)
    if not items:
        st.error("❌ Статьи не найдены.")
        return

    n_analyzed = len(items)
    st.success(f"📄 Найдено анализируемых статей: {n_analyzed}")
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Валидация данных
    overall_status.text("🔍 Валидация данных...")
    validated_items = validate_and_clean_data(items)
    journal_prefix = get_doi_prefix(validated_items[0].get('DOI', '')) if validated_items else ''
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Обработка анализируемых статей
    overall_status.text("🔄 Обработка анализируемых статей...")
    st.info("🔄 Обработка анализируемых статей...")
    
    analyzed_metadata = []
    dois = [item.get('DOI') for item in validated_items if item.get('DOI')]
    
    meta_progress = st.progress(0)
    for i, doi in enumerate(dois):
        result = get_unified_metadata(doi)
        analyzed_metadata.append({
            'doi': doi,
            'crossref': result['crossref'],
            'openalex': result['openalex']
        })
        meta_progress.progress((i + 1) / len(dois))
    
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Получение цитирующих работ
    overall_status.text("🔗 Сбор цитирующих работ...")
    st.info("🔗 Сбор цитирующих работ...")
    
    all_citing_metadata = []
    analyzed_dois = [am['doi'] for am in analyzed_metadata if am.get('doi')]
    
    citing_progress = st.progress(0)
    for i, doi in enumerate(analyzed_dois):
        citings = get_citing_dois_and_metadata(doi)
        all_citing_metadata.extend(citings)
        citing_progress.progress((i + 1) / len(analyzed_dois))
    
    n_citing = len(set(c['doi'] for c in all_citing_metadata if c.get('doi')))
    st.success(f"📄 Уникальных цитирующих работ: {n_citing}")
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Расчет статистики
    overall_status.text("📊 Расчет статистики...")
    st.info("📊 Расчет статистики...")
    
    analyzed_stats = extract_stats_from_metadata(analyzed_metadata, journal_prefix=journal_prefix)
    citing_stats = extract_stats_from_metadata(all_citing_metadata, is_analyzed=False)
    enhanced_stats = enhanced_stats_calculation(analyzed_metadata, all_citing_metadata)
    
    overlap_stats = analyze_overlaps(analyzed_metadata, all_citing_metadata)
    
    current_date = datetime.now()
    if_days = calculate_if_and_days(analyzed_metadata, all_citing_metadata, current_date)
    
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Создание отчета
    overall_status.text("💾 Создание отчета...")
    st.info("💾 Создание Excel отчета...")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'journal_analysis_{issn}_{timestamp}.xlsx'
    excel_buffer = create_enhanced_excel_report(analyzed_metadata, all_citing_metadata, analyzed_stats, citing_stats, enhanced_stats, if_days, overlap_stats, filename)
    
    current_step += 1
    overall_progress.progress(current_step / steps)

    # Вывод результатов в интерфейс
    st.markdown("---")
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric("Impact Factor", f"{if_days['if_value']:.4f}")
        st.metric("H-index", enhanced_stats['h_index'])
        st.metric("Всего статей", analyzed_stats['n_items'])
        st.metric("Всего цитирований", enhanced_stats['total_citations'])
    
    with col2:
        st.metric("Среднее цитирований на статью", f"{enhanced_stats['avg_citations_per_article']:.1f}")
        st.metric("Статьи с ≥10 цитированиями", analyzed_stats['articles_with_10_citations'])
        st.metric("Статьи с ≥50 цитированиями", analyzed_stats['articles_with_50_citations'])
        st.metric("Статьи с ≥100 цитированиями", analyzed_stats['articles_with_100_citations'])
    
    st.subheader("🔀 Пересечения между анализируемыми и цитирующими")
    overlap_df = pd.DataFrame({
        'Категория': ['Авторы', 'Аффилиации', 'Страны'],
        'Количество': [overlap_stats['author_overlap_count'], overlap_stats['affiliation_overlap_count'], overlap_stats['country_overlap_count']],
        'Процент': [f"{overlap_stats['author_overlap_pct']:.1f}%", f"{overlap_stats['affiliation_overlap_pct']:.1f}%", f"{overlap_stats['country_overlap_pct']:.1f}%"]
    })
    st.dataframe(overlap_df)
    
    # Детали пересечений
    if overlap_stats['overlap_pairs']:
        pairs_df = pd.DataFrame(overlap_stats['overlap_pairs'])
        st.dataframe(pairs_df.head(10))  # Show top 10
        st.caption(f"Всего пар с пересечениями: {len(overlap_stats['overlap_pairs'])}")
    else:
        st.info("Нет пересечений.")
    
    st.subheader("👥 Авторы и организации")
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Анализируемые: Всего авторов", len(analyzed_stats['all_authors_list']))
        st.metric("Анализируемые: Уникальных стран", analyzed_stats['unique_countries_count'])
    with col2:
        st.metric("Цитирующие: Всего авторов", len(citing_stats['all_authors_list']))
        st.metric("Цитирующие: Уникальных стран", citing_stats['unique_countries_count'])
    
    st.subheader("📊 Журналы цитирующих работ")
    st.metric("Уникальных журналов", citing_stats['unique_journals_count'])
    st.metric("Уникальных издателей", citing_stats['unique_publishers_count'])
    
    st.subheader("⏱️ Лаги цитирования")
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Min дни до первого", if_days['first_citation']['min_days_to_first_citation'])
        st.metric("Max дни до первого", if_days['first_citation']['max_days_to_first_citation'])
        st.metric("Mean дни до первого", f"{if_days['first_citation']['mean_days_to_first_citation']:.1f}")
    with col2:
        st.metric("Min лаг всех", if_days['all_lags']['min_lag_days'])
        st.metric("Max лаг всех", if_days['all_lags']['max_lag_days'])
        st.metric("Mean лаг всех", f"{if_days['all_lags']['mean_lag_days']:.1f}")
    
    st.subheader("📈 Скорость накопления цитирований")
    st.metric("Лет покрыто", if_days['total_years_covered'])
    
    # Визуализации
    st.subheader("📊 Визуализации")
    
    # Гистограмма цитирований по годам
    if if_days['yearly_citations']:
        yearly_df = pd.DataFrame(if_days['yearly_citations'])
        fig_yearly = px.bar(yearly_df, x='year', y='citations_count', title="Цитирования по годам")
        st.plotly_chart(fig_yearly, use_container_width=True)
    
    # Топ авторы
    top_authors_analyzed = pd.DataFrame(analyzed_stats['all_authors'][:10])
    if not top_authors_analyzed.empty:
        fig_auth = px.bar(top_authors_analyzed, x='Количество статей', y='Автор', orientation='h', title="Топ авторы (анализируемые)")
        st.plotly_chart(fig_auth, use_container_width=True)
    
    # Топ страны
    top_countries = pd.DataFrame(citing_stats['all_countries'][:10])
    if not top_countries.empty:
        fig_countries = px.pie(top_countries, values='Количество упоминаний', names='Страна', title="Топ страны (цитирующие)")
        st.plotly_chart(fig_countries, use_container_width=True)
    
    # Кривые накопления
    accumulation_data = []
    for pub_year, curve in if_days['accumulation_curves'].items():
        for point in curve:
            accumulation_data.append({
                'Pub Year': pub_year,
                'Years Since': point['years_since_publication'],
                'Cumulative': point['cumulative_citations']
            })
    if accumulation_data:
        acc_df = pd.DataFrame(accumulation_data)
        fig_acc = px.line(acc_df, x='Years Since', y='Cumulative', color='Pub Year', title="Кривые накопления цитирований")
        st.plotly_chart(fig_acc, use_container_width=True)
    
    # Сеть цитирований
    if enhanced_stats['citation_network']:
        net_data = []
        for year, years in enhanced_stats['citation_network'].items():
            for y in years:
                net_data.append({'Pub Year': year, 'Cite Year': y})
        net_df = pd.DataFrame(net_data)
        fig_net = px.scatter(net_df, x='Pub Year', y='Cite Year', title="Сеть цитирований")
        st.plotly_chart(fig_net, use_container_width=True)
    
    # Download
    st.download_button(
        label="💾 Скачать Excel отчет",
        data=excel_buffer.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.success("✅ Анализ завершен!")

# === Streamlit Interface ===
st.set_page_config(page_title="Анализ Журнала", page_icon="🔬", layout="wide")

st.title("🔬 АНАЛИЗ ЖУРНАЛА: ПОЛНАЯ СТАТИСТИКА")
st.markdown("---")

st.sidebar.header("📝 Введите данные")
issn = st.sidebar.text_input("ISSN", value='2411-1414', help="Введите ISSN журнала")
period = st.sidebar.text_input("Период", value='2022-2024', help="Примеры: 2022, 2022-2025, 2022,2024")

# Автоопределение названия
if issn:
    with st.spinner("Определение названия журнала..."):
        journal_name = get_journal_name(issn)
        st.sidebar.info(f"📖 Журнал: {journal_name}")

email = st.sidebar.text_input("Email (для API)", value=EMAIL, help="Для Crossref API")

if st.sidebar.button("🚀 Запустить анализ"):
    if not issn or not period:
        st.error("Введите ISSN и период!")
    else:
        analyze_journal(issn, period)

st.sidebar.markdown("---")
st.sidebar.info("""
### 💡 Инструкция:
- Замените EMAIL на реальный (используйте secrets.toml для продакшена)
- Анализ может занять несколько минут
- Результаты: визуализации + Excel
""")

st.sidebar.markdown("### 🔍 Примеры ISSN:")
test_journals = [
    ("1367-4803", "Bioinformatics"),
    ("0028-0836", "Nature"), 
    ("1476-4687", "Nature Communications"),
    ("0001-6772", "Acta Physiologica"),
    ("2411-1414", "Journal of Physics: Conference Series")
]

for issn_ex, name in test_journals:
    st.sidebar.markdown(f"• {issn_ex} - {name}")



